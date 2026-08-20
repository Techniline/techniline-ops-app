"""
Import SKU data from brand volume charts into packing_sku_catalog.

Usage:
  python scripts/import-packing-skus.py

Reads 4 Excel files and outputs SQL to stdout.
Paste the output into Supabase → SQL Editor.

Requirements: pip install openpyxl
"""

import sys
import os
import math
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

FILES = {
    "wharfedale": r"C:\Users\vihan\Downloads\Wharfedale Volume Chart_140525 1.xlsx",
    "behringer":  r"C:\Users\vihan\Downloads\Behringer Volume Chart_08.06.26 1.xlsx",
    "ahuja":      r"C:\Users\vihan\Downloads\Ahuja Volume Chart_140525 1.xlsx",
    "hikvision":  r"C:\Users\vihan\Downloads\Volume Chart - Hikvision 2.xlsx",
}

def safe_num(v):
    """Return float or None."""
    if v is None or v == "" or v == "#N/A":
        return None
    try:
        f = float(str(v).replace(",", ""))
        return f if math.isfinite(f) else None
    except (ValueError, TypeError):
        return None

def safe_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def sql_val(v):
    if v is None: return "NULL"
    if isinstance(v, (int, float)): return str(v)
    return "'" + str(v).replace("'", "''") + "'"

rows = []

# ─── Wharfedale ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILES["wharfedale"], data_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
# Item Code, Description, GW, Volume, HS Code
idx = {h: i for i, h in enumerate(headers) if h}
for row in ws.iter_rows(min_row=2, values_only=True):
    model_no = safe_str(row[idx.get("Item Code", 0)])
    if not model_no: continue
    rows.append({
        "model_no": model_no,
        "brand": "Wharfedale",
        "description": safe_str(row[idx.get("Description", 1)]),
        "hs_code": safe_str(row[idx.get("HS Code", 4)]),
        "country_of_origin": "China",
        "unit_weight_kg": safe_num(row[idx.get("GW", 2)]),
        "unit_cbm": safe_num(row[idx.get("Volume", 3)]),
        "carton_qty": None,
        "carton_weight_kg": None,
        "carton_cbm": None,
    })

# ─── Behringer ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILES["behringer"], data_only=True)
ws = wb.active
headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
col = {h: i for i, h in enumerate(headers) if h}
for row in ws.iter_rows(min_row=2, values_only=True):
    model_no = safe_str(row[col.get("TLE SKU", 2)])
    if not model_no: continue
    l = safe_num(row[col.get("Single Unit Length (mm)", 10)])
    w = safe_num(row[col.get("Single Unit Width (mm)", 11)])
    h_ = safe_num(row[col.get("Single Unit Height (mm)", 12)])
    unit_cbm = round(l * w * h_ / 1_000_000_000, 6) if l and w and h_ else None
    rows.append({
        "model_no": model_no,
        "brand": "Behringer",
        "description": safe_str(row[col.get("Headline", 3)]),
        "hs_code": None,
        "country_of_origin": "China",
        "unit_weight_kg": safe_num(row[col.get("Single Unit Weight (kg)", 13)]),
        "unit_cbm": unit_cbm,
        "carton_qty": int(safe_num(row[col.get("Master Carton Packaging Qty", 18)]) or 0) or None,
        "carton_weight_kg": safe_num(row[col.get("Master Carton Gross Weight (kg)", 19)]),
        "carton_cbm": safe_num(row[col.get("Master Carton Volume (CBM)", 17)]),
    })

# ─── Ahuja ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILES["ahuja"], data_only=True)
ws = wb.active
headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
col = {h: i for i, h in enumerate(headers) if h}
for row in ws.iter_rows(min_row=2, values_only=True):
    model_no = safe_str(row[col.get("Item Code", 0)])
    if not model_no: continue
    rows.append({
        "model_no": model_no,
        "brand": "Ahuja",
        "description": safe_str(row[col.get("Description", 1)]),
        "hs_code": safe_str(row[col.get("HS Code", 9)]),
        "country_of_origin": "India",
        "unit_weight_kg": None,
        "unit_cbm": safe_num(row[col.get("Volume per pc", 8)]),
        "carton_qty": int(safe_num(row[col.get("Master Carton Packaging Qty", 7)]) or 0) or None,
        "carton_weight_kg": safe_num(row[col.get("Master Carton GW", 6)]),
        "carton_cbm": safe_num(row[col.get("Master Carton Volume", 5)]),
    })

# ─── Hikvision ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILES["hikvision"], data_only=True)
ws = wb.active
headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
col = {h: i for i, h in enumerate(headers) if h}
for row in ws.iter_rows(min_row=2, values_only=True):
    model_no = safe_str(row[col.get("TLE SKU", 5)])
    if not model_no or model_no == "#N/A": continue
    uw_g = safe_num(row[col.get("Packing for One Weight(g)", 11)])
    cw_g = safe_num(row[col.get("Packing for Multi Weight(g)", 17)])
    rows.append({
        "model_no": model_no,
        "brand": "Hikvision",
        "description": safe_str(row[col.get("Functional Description", 3)]),
        "hs_code": None,
        "country_of_origin": "China",
        "unit_weight_kg": round(uw_g / 1000, 4) if uw_g else None,
        "unit_cbm": safe_num(row[col.get("CBM/pc", 10)]),
        "carton_qty": int(safe_num(row[col.get("MOQ", 19)]) or 0) or None,
        "carton_weight_kg": round(cw_g / 1000, 4) if cw_g else None,
        "carton_cbm": safe_num(row[col.get("CBM/Master Carton", 16)]),
    })

# ─── Output SQL ────────────────────────────────────────────────────────────
print(f"-- Importing {len(rows)} SKUs into packing_sku_catalog")
print("-- Run in Supabase → SQL Editor")
print()

CHUNK = 100
for i in range(0, len(rows), CHUNK):
    chunk = rows[i:i+CHUNK]
    vals = []
    for r in chunk:
        vals.append(
            f"({sql_val(r['model_no'])}, {sql_val(r['brand'])}, {sql_val(r['description'])}, "
            f"{sql_val(r['hs_code'])}, {sql_val(r['country_of_origin'])}, "
            f"{sql_val(r['unit_weight_kg'])}, {sql_val(r['unit_cbm'])}, "
            f"{sql_val(r['carton_qty'])}, {sql_val(r['carton_weight_kg'])}, {sql_val(r['carton_cbm'])}, "
            f"'import')"
        )
    print(
        "INSERT INTO public.packing_sku_catalog "
        "(model_no, brand, description, hs_code, country_of_origin, "
        "unit_weight_kg, unit_cbm, carton_qty, carton_weight_kg, carton_cbm, source)\n"
        "VALUES\n  " + ",\n  ".join(vals) + "\n"
        "ON CONFLICT (model_no) DO UPDATE SET\n"
        "  brand = EXCLUDED.brand,\n"
        "  description = COALESCE(EXCLUDED.description, packing_sku_catalog.description),\n"
        "  hs_code = COALESCE(EXCLUDED.hs_code, packing_sku_catalog.hs_code),\n"
        "  country_of_origin = EXCLUDED.country_of_origin,\n"
        "  unit_weight_kg = COALESCE(EXCLUDED.unit_weight_kg, packing_sku_catalog.unit_weight_kg),\n"
        "  unit_cbm = COALESCE(EXCLUDED.unit_cbm, packing_sku_catalog.unit_cbm),\n"
        "  carton_qty = COALESCE(EXCLUDED.carton_qty, packing_sku_catalog.carton_qty),\n"
        "  carton_weight_kg = COALESCE(EXCLUDED.carton_weight_kg, packing_sku_catalog.carton_weight_kg),\n"
        "  carton_cbm = COALESCE(EXCLUDED.carton_cbm, packing_sku_catalog.carton_cbm),\n"
        "  updated_at = now();\n"
    )

print(f"-- Done: {len(rows)} rows processed")
